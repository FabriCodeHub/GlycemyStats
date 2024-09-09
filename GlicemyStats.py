import csv
from datetime import datetime
import pyfiglet
import openpyxl
from openpyxl.styles import PatternFill, Font
from openpyxl.chart import LineChart, Reference
import re
import os
import statistics

def stampa_titolo():
    titolo = pyfiglet.figlet_format("GlycemyStats")
    print(titolo)

def valida_data(data):
    formati = ["%Y-%m-%d", "%d/%m/%Y", "%d\\%m\\%Y"]
    for formato in formati:
        try:
            return datetime.strptime(data, formato).date()
        except ValueError:
            pass
    raise ValueError("Formato data non valido")

def valida_ora(ora):
    formati = ["%H:%M", "%H.%M"]
    for formato in formati:
        try:
            return datetime.strptime(ora, formato).time()
        except ValueError:
            pass
    raise ValueError("Formato ora non valido")

def inserisci_dati_glicemia():
    while True:
        try:
            data = input("Inserisci la data (YYYY-MM-DD o DD/MM/YYYY o DD\\MM\\YYYY): ")
            data_valida = valida_data(data)
            
            ora = input("Inserisci l'ora (HH:MM o HH.MM): ")
            ora_valida = valida_ora(ora)
            
            livello = input("Inserisci il livello di glicemia (2 o 3 cifre): ")
            if not re.match(r"^\d{2,3}$", livello):
                raise ValueError("Livello di glicemia non valido")
            livello = float(livello)
            
            break
        except ValueError as ve:
            print(f"Errore: {ve}. Riprova.")
    
    data_ora = datetime.combine(data_valida, ora_valida)
    
    with open("livelli_glicemia.csv", mode='a', newline='') as file:
        writer = csv.writer(file)
        writer.writerow([data_ora.strftime("%Y-%m-%d"), data_ora.strftime("%H:%M"), livello])
    print("Dati inseriti correttamente.")

def crea_file_excel():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Livelli di Glicemia"
    
    ws.append(["Data", "Ora", "Livello di Glicemia"])
    
    with open("livelli_glicemia.csv", mode='r') as file:
        reader = csv.reader(file)
        for row in reader:
            ws.append(row)
    
    applica_formattazione(ws)
    crea_grafico(ws)
    aggiungi_statistiche(ws)
    
    wb.save("livelli_glicemia.xlsx")

def applica_formattazione(ws):
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    
    for row in ws.iter_rows(min_row=2, min_col=3, max_col=3):
        for cell in row:
            value = cell.value
            if value > 180:
                cell.fill = red_fill
            elif 140 <= value <= 180:
                cell.fill = yellow_fill
            elif 70 <= value <= 139:
                cell.fill = green_fill
            else:
                cell.fill = red_fill

def crea_grafico(ws):
    chart = LineChart()
    chart.title = "Livelli di Glicemia nel Tempo"
    chart.style = 10
    chart.y_axis.title = 'Livello di Glicemia'
    chart.x_axis.title = 'Data e Ora'
    
    data = Reference(ws, min_col=3, min_row=1, max_row=ws.max_row)
    categories = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)
    
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)
    
    ws.add_chart(chart, "E5")

def aggiungi_statistiche(ws):
    livelli = [cell.value for row in ws.iter_rows(min_row=2, min_col=3, max_col=3) for cell in row]
    
    media = statistics.mean(livelli)
    mediana = statistics.median(livelli)
    deviazione_std = statistics.stdev(livelli)
    
    ws['A{}'.format(ws.max_row + 2)] = "Statistiche"
    ws['A{}'.format(ws.max_row + 1)] = "Media"
    ws['B{}'.format(ws.max_row)] = media
    ws['A{}'.format(ws.max_row + 1)] = "Mediana"
    ws['B{}'.format(ws.max_row)] = mediana
    ws['A{}'.format(ws.max_row + 1)] = "Deviazione Standard"
    ws['B{}'.format(ws.max_row)] = deviazione_std

def visualizza_statistiche():
    if not os.path.exists("livelli_glicemia.csv"):
        print("Nessun dato disponibile.")
        return
    
    with open("livelli_glicemia.csv", mode='r') as file:
        reader = csv.reader(file)
        livelli = [float(row[2]) for row in reader]
    
    if not livelli:
        print("Nessun dato disponibile.")
        return
    
    print("\nStatistiche dei livelli di glicemia:")
    print(f"Media: {statistics.mean(livelli):.2f}")
    print(f"Mediana: {statistics.median(livelli):.2f}")
    print(f"Deviazione Standard: {statistics.stdev(livelli):.2f}")
    print(f"Valore Minimo: {min(livelli):.2f}")
    print(f"Valore Massimo: {max(livelli):.2f}")

def main():
    stampa_titolo()
    while True:
        print("\nMenu:")
        print("1. Inserisci dati glicemia")
        print("2. Crea file Excel")
        print("3. Visualizza statistiche")
        print("4. Esci")
        
        scelta = input("Scegli un'opzione (1-4): ")
        
        if scelta == '1':
            inserisci_dati_glicemia()
        elif scelta == '2':
            crea_file_excel()
            print("File Excel creato correttamente: livelli_glicemia.xlsx")
        elif scelta == '3':
            visualizza_statistiche()
        elif scelta == '4':
            print("Grazie per aver usato GlycemyStats!")
            break
        else:
            print("Opzione non valida. Riprova.")

if __name__ == "__main__":
    main()
